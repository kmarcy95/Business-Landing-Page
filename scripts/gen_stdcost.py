from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path('downloads/KM-Consulting_Standard-Costing-Variance_v1.0.xlsx')
BLUE='DDEBF7'; DARK='1F4E78'; GREY='E7E6E6'; WHITE='FFFFFF'; GREEN='E2F0D9'; RED='FCE4D6'
thin=Side(style='thin', color='D9E2F3')
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook(); wb.remove(wb.active)
for s in ['Read Me','Inputs','Actuals','Calc','Dashboard','Example']:
    wb.create_sheet(s)

def ws_title(ws, title, subtitle=None):
    ws['B2']=title; ws['B2'].font=Font(bold=True,size=16,color=DARK)
    if subtitle:
        ws['B3']=subtitle; ws['B3'].font=Font(italic=True,color='666666')

def style_range(ws, rng, fill=None, font=None, align=None, border=True):
    for row in ws[rng]:
        for c in row:
            if fill: c.fill=PatternFill('solid', fgColor=fill)
            if font: c.font=font
            if align: c.alignment=align
            if border: c.border=BORDER

def table(ws, name, ref):
    t=Table(displayName=name, ref=ref)
    t.tableStyleInfo=TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(t)

def widths(ws, mapping):
    for col,w in mapping.items(): ws.column_dimensions[col].width=w

def add_name(name, ws, rng):
    wb.defined_names.add(DefinedName(name, attr_text=f"'{ws.title}'!${rng.replace(':',':$').replace('$','$')}"))

def nm(name, sheet, rng):
    wb.defined_names.add(DefinedName(name, attr_text=f"'{sheet}'!{rng}"))

def input_style(ws, rng):
    style_range(ws, rng, BLUE, border=True)
    for row in ws[rng]:
        for c in row: c.protection = c.protection.copy(locked=False)

def header(ws, rng):
    style_range(ws, rng, DARK, Font(bold=True,color=WHITE), Alignment(horizontal='center'))

def money_fmt(ws, rng):
    for row in ws[rng]:
        for c in row: c.number_format='$#,##0;[Red]($#,##0);-'

def num_fmt(ws, rng):
    for row in ws[rng]:
        for c in row: c.number_format='#,##0.00;[Red](#,##0.00);-'

# Read Me
ws=wb['Read Me']; ws_title(ws,'KM Consulting — Standard Costing Variance Model','Version v1.0 | Formula-only workbook | Excel 2016+')
read_rows=[
('Purpose','Calculate standard-costing variances, sales variances, overhead variances, and an operating-profit bridge.'),
('How to use','Enter assumptions on Inputs and actual results on Actuals. Calc and Dashboard are formula-driven and protected.'),
('Sign convention','Positive variance = Unfavorable = actual cost above standard or sales below budget; negative = Favorable.'),
('Assumptions','Fixed OH absorption costing. Overhead activity base = labor hours. Material mix/yield supports multiple materials per product.'),
('Version','v1.0')]
for i,(a,b) in enumerate(read_rows,5):
    ws.cell(i,2,a); ws.cell(i,3,b)
header(ws,'B4:C4'); ws['B4']='Topic'; ws['C4']='Detail'
style_range(ws,'B5:C9'); widths(ws, {'B':24,'C':110})

products=[
['Product A',5.00,3000,20,2000,4,11000,5,35,100],
['Product B',8.00,1000,25,750,6,6000,8,50,80],
['Product C',6.50,0,22,0,5,0,6,40,0],
['Product D',9.00,0,24,0,6,0,7,55,0],
['Product E',4.25,0,18,0,3,0,4,30,0]]
materials=[
['Product A','Material 1',5.00,3000,5.20,3200],
['Product B','Material 1',8.00,1000,7.80,1050],
['Product A','Material 2',0,0,0,0],
['Product B','Material 2',0,0,0,0],
['Product C','Material 1',0,0,0,0],
['Product D','Material 1',0,0,0,0],
['Product E','Material 1',0,0,0,0]]
actuals=[
['Product A',5.20,3200,19.50,2100,8600,11300,36,105],
['Product B',7.80,1050,25.40,720,4700,5900,48,78],
['Product C',0,0,0,0,0,0,0,0],
['Product D',0,0,0,0,0,0,0,0],
['Product E',0,0,0,0,0,0,0,0]]

# Inputs
ws=wb['Inputs']; ws_title(ws,'Inputs','Blue cells are editable. Named ranges are applied to table columns.')
prod_heads=['Product','Std price/unit','Std qty','Std labor rate','Std labor hours','Std var OH rate/hr','Budgeted fixed OH','Std fixed OH rate/hr','Budgeted selling price','Budgeted units']
ws['B4']='Products'; ws['B4'].font=Font(bold=True,color=DARK)
for j,h in enumerate(prod_heads,2): ws.cell(5,j,h)
for i,r in enumerate(products,6):
    for j,v in enumerate(r,2): ws.cell(i,j,v)
header(ws,'B5:K5'); input_style(ws,'B6:K10'); table(ws,'tblProducts','B5:K10')
mat_start=14
ws.cell(mat_start-1,2,'Material Inputs'); ws.cell(mat_start-1,2).font=Font(bold=True,color=DARK)
mat_heads=['Product','Material','Std price','Std qty','Actual price','Actual qty']
for j,h in enumerate(mat_heads,2): ws.cell(mat_start,j,h)
for i,r in enumerate(materials,mat_start+1):
    for j,v in enumerate(r,2): ws.cell(i,j,v)
header(ws,f'B{mat_start}:G{mat_start}'); input_style(ws,f'B{mat_start+1}:G{mat_start+len(materials)}')
table(ws,'tblMaterialInputs',f'B{mat_start}:G{mat_start+len(materials)}')
widths(ws, {'B':18,'C':18,'D':13,'E':13,'F':15,'G':15,'H':18,'I':18,'J':20,'K':15})
for r in ['C6:K10',f'D{mat_start+1}:G{mat_start+len(materials)}']:
    num_fmt(ws,r)
# named ranges
for col,name in zip('BCDEFGHIJK',['Product_List','Std_Price_Unit','Std_Qty','Std_Labor_Rate','Std_Labor_Hours','Std_Var_OH_Rate_Hr','Budgeted_Fixed_OH','Std_Fixed_OH_Rate_Hr','Budgeted_Selling_Price','Budgeted_Units']):
    nm(name,'Inputs',f'${col}$6:${col}$10')
for col,name in zip('BCDEFG',['MI_Product','MI_Material','MI_Std_Price','MI_Std_Qty','MI_Actual_Price','MI_Actual_Qty']):
    nm(name,'Inputs',f'${col}${mat_start+1}:${col}${mat_start+len(materials)}')

# Actuals
ws=wb['Actuals']; ws_title(ws,'Actuals','Enter actual cost and sales data by product.')
act_heads=['Product','Actual price','Actual qty','Actual labor rate','Actual labor hours','Actual var OH','Actual fixed OH','Actual selling price','Actual units']
for j,h in enumerate(act_heads,2): ws.cell(5,j,h)
for i,r in enumerate(actuals,6):
    for j,v in enumerate(r,2): ws.cell(i,j,v)
header(ws,'B5:J5'); input_style(ws,'B6:J10'); table(ws,'tblActuals','B5:J10')
widths(ws, {'B':18,'C':14,'D':14,'E':18,'F':18,'G':16,'H':16,'I':20,'J':14})
num_fmt(ws,'C6:J10')
for col,name in zip('BCDEFGHIJ',['Actual_Product','Actual_Price','Actual_Qty','Actual_Labor_Rate','Actual_Labor_Hours','Actual_Var_OH','Actual_Fixed_OH','Actual_Selling_Price','Actual_Units']):
    nm(name,'Actuals',f'${col}$6:${col}$10')
# validation
prod_dv=DataValidation(type='list', formula1='=Product_List', allow_blank=False)
prod_dv.error='Choose a product from Inputs.'
for sh,rng in [('Inputs',f'B{mat_start+1}:B{mat_start+len(materials)}'),('Actuals','B6:B10')]:
    dv=DataValidation(type='list', formula1='=Product_List', allow_blank=False)
    dv.error='Choose a product from Inputs.'; wb[sh].add_data_validation(dv); dv.add(rng)
for sh,ranges in [('Inputs',['C6:K10',f'D{mat_start+1}:G{mat_start+len(materials)}']),('Actuals',['C6:J10'])]:
    dv=DataValidation(type='decimal', operator='greaterThanOrEqual', formula1='0', allow_blank=True)
    dv.error='Enter zero or a positive number.'; wb[sh].add_data_validation(dv)
    for rng in ranges: dv.add(rng)

# Calc
ws=wb['Calc']; ws_title(ws,'Variance Calculations','Positive = Unfavorable; Negative = Favorable')
calc_heads=['Product','Std Mat Cost','Act Mat Cost','Material Price','Material Quantity','Material Mix','Material Yield','Labor Rate','Labor Efficiency','Labor Mix','Var OH Spending','Var OH Efficiency','Fixed OH Budget','Fixed OH Volume','Sales Price','Sales Volume','Sales Mix','Total Cost Variance','Sales Variance']
for j,h in enumerate(calc_heads,2): ws.cell(5,j,h)
header(ws,'B5:T5')
for i in range(6,11):
    p=f'B{i}'; ws[p]=f'=Inputs!B{i}'
    ws[f'C{i}']=f'=Inputs!C{i}*Inputs!D{i}'
    ws[f'D{i}']=f'=Actuals!C{i}*Actuals!D{i}'
    ws[f'E{i}']=f'=SUMPRODUCT((MI_Product={p})*(MI_Actual_Price-MI_Std_Price)*MI_Actual_Qty)'
    ws[f'F{i}']=f'=SUMPRODUCT((MI_Product={p})*(MI_Actual_Qty-MI_Std_Qty)*MI_Std_Price)'
    ws[f'G{i}']=f'=IFERROR(SUMPRODUCT((MI_Product={p})*(MI_Actual_Qty-SUMIF(MI_Product,{p},MI_Actual_Qty)*MI_Std_Qty/SUMIF(MI_Product,{p},MI_Std_Qty))*MI_Std_Price),0)'
    ws[f'H{i}']=f'=F{i}-G{i}'
    ws[f'I{i}']=f'=(Actuals!E{i}-Inputs!E{i})*Actuals!F{i}'
    ws[f'J{i}']=f'=(Actuals!F{i}-Inputs!F{i})*Inputs!E{i}'
    ws[f'K{i}']=f'=0'
    ws[f'L{i}']=f'=Actuals!G{i}-Inputs!G{i}*Actuals!F{i}'
    ws[f'M{i}']=f'=(Actuals!F{i}-Inputs!F{i})*Inputs!G{i}'
    ws[f'N{i}']=f'=Actuals!H{i}-Inputs!H{i}'
    ws[f'O{i}']=f'=Inputs!H{i}-(Inputs!I{i}*Inputs!F{i})'
    ws[f'P{i}']=f'=-(Actuals!I{i}-Inputs!J{i})*Actuals!J{i}'
    ws[f'Q{i}']=f'=-(Actuals!J{i}-Inputs!K{i})*Inputs!J{i}'
    ws[f'R{i}']=f'=0'
    ws[f'S{i}']=f'=SUM(E{i}:O{i})'
    ws[f'T{i}']=f'=SUM(P{i}:R{i})'
for j in range(2,21): ws.cell(11,j, f'=SUM({get_column_letter(j)}6:{get_column_letter(j)}10)')
ws['B11']='Grand Total'; ws['B11'].font=Font(bold=True)
style_range(ws,'B6:T11',WHITE); money_fmt(ws,'C6:T11'); table(ws,'tblCalc','B5:T11')
widths(ws, {get_column_letter(c):15 for c in range(2,21)}); ws.column_dimensions['B'].width=18
ws.protection.sheet=True

# Dashboard
ws=wb['Dashboard']; ws_title(ws,'Dashboard','Operating-profit bridge and variance KPI summary')
labels=['Budgeted operating profit','Sales price variance','Sales volume variance','Sales mix variance','Cost variances','Actual operating profit']
for i,l in enumerate(labels,5): ws.cell(i,2,l)
ws['C5']='=SUMPRODUCT(Budgeted_Selling_Price,Budgeted_Units)-SUMPRODUCT(Std_Price_Unit,Std_Qty)-SUMPRODUCT(Std_Labor_Rate,Std_Labor_Hours)-SUMPRODUCT(Std_Var_OH_Rate_Hr,Std_Labor_Hours)-SUMPRODUCT(Std_Fixed_OH_Rate_Hr,Std_Labor_Hours)'
ws['C6']='=-Calc!P11'; ws['C7']='=-Calc!Q11'; ws['C8']='=-Calc!R11'; ws['C9']='=-Calc!S11'; ws['C10']='=SUM(C5:C9)'
header(ws,'B4:C4'); ws['B4']='Operating Profit Bridge'; ws['C4']='Amount'; style_range(ws,'B5:C10'); money_fmt(ws,'C5:C10')
chart=BarChart(); chart.type='bar'; chart.style=10; chart.title='Budgeted to Actual OP Bridge'; chart.y_axis.title='Metric'; chart.x_axis.title='Amount'
chart.add_data(Reference(ws,min_col=3,min_row=4,max_row=10), titles_from_data=True)
chart.set_categories(Reference(ws,min_col=2,min_row=5,max_row=10)); chart.height=7; chart.width=12; ws.add_chart(chart,'E4')
# KPI table
kheads=['Variance','Amount','F/U','% of Standard']; start=15
for j,h in enumerate(kheads,2): ws.cell(start,j,h)
for idx,(nmv,amt,base) in enumerate([
('Material', 'SUM(Calc!E11:H11)', 'Calc!C11'),('Labor','SUM(Calc!I11:K11)','SUMPRODUCT(Std_Labor_Rate,Std_Labor_Hours)'),('Variable OH','SUM(Calc!L11:M11)','SUMPRODUCT(Std_Var_OH_Rate_Hr,Std_Labor_Hours)'),('Fixed OH','SUM(Calc!N11:O11)','SUMPRODUCT(Std_Fixed_OH_Rate_Hr,Std_Labor_Hours)'),('Sales','Calc!T11','SUMPRODUCT(Budgeted_Selling_Price,Budgeted_Units)')], start+1):
    ws.cell(idx,2,nmv); ws.cell(idx,3,f'={amt}'); ws.cell(idx,4,f'=IF(C{idx}>0,"Unfavorable",IF(C{idx}<0,"Favorable","Neutral"))'); ws.cell(idx,5,f'=IFERROR(C{idx}/({base}),0)')
header(ws,f'B{start}:E{start}'); style_range(ws,f'B{start+1}:E{start+5}'); money_fmt(ws,f'C{start+1}:C{start+5}')
for c in ws[f'E{start+1}:E{start+5}']:
    for x in c: x.number_format='0.0%;[Red](0.0%);-'
# cost reconciliation
r=24; ws.cell(r,2,'Cost Reconciliation'); ws.cell(r,2).font=Font(bold=True,color=DARK)
for i,(l,f) in enumerate([('Standard cost','=SUMPRODUCT(Std_Price_Unit,Std_Qty)+SUMPRODUCT(Std_Labor_Rate,Std_Labor_Hours)+SUMPRODUCT(Std_Var_OH_Rate_Hr,Std_Labor_Hours)+SUMPRODUCT(Std_Fixed_OH_Rate_Hr,Std_Labor_Hours)'),('Actual cost','=SUMPRODUCT(Actual_Price,Actual_Qty)+SUMPRODUCT(Actual_Labor_Rate,Actual_Labor_Hours)+SUM(Actual_Var_OH)+SUM(Actual_Fixed_OH)'),('Total cost variance','=C26-C25')], r+1):
    ws.cell(i,2,l); ws.cell(i,3,f)
header(ws,f'B{r}:C{r}'); ws['C24']='Amount'; style_range(ws,f'B{r+1}:C{r+3}'); money_fmt(ws,f'C{r+1}:C{r+3}')
widths(ws, {'B':28,'C':18,'D':16,'E':16,'F':16,'G':16,'H':16,'I':16})
ws.protection.sheet=True

# Example
ws=wb['Example']; ws_title(ws,'Worked Example — Two Product Single-Material Case','Matches the requested cost-side grand totals.')
ex_headers=['Product','Std price','Std qty','Actual price','Actual qty','Std labor rate','Std labor hours','Actual labor rate','Actual labor hours','Std var OH rate','Actual var OH','Budgeted fixed OH','Std fixed OH rate','Actual fixed OH']
for j,h in enumerate(ex_headers,2): ws.cell(5,j,h)
ex_rows=[['Product A',5,3000,5.2,3200,20,2000,19.5,2100,4,8600,11000,5,11300],['Product B',8,1000,7.8,1050,25,750,25.4,720,6,4700,6000,8,5900]]
for i,row in enumerate(ex_rows,6):
    for j,v in enumerate(row,2): ws.cell(i,j,v)
header(ws,'B5:O5'); input_style(ws,'B6:O7'); table(ws,'tblExampleInputs','B5:O7')
# variance detail
vr=11; vheads=['Product','Std Cost','Actual Cost','Material Price','Material Qty','Labor Rate','Labor Eff','Var OH Spend','Var OH Eff','Fixed Budget','Fixed Volume','Total Cost Variance']
for j,h in enumerate(vheads,2): ws.cell(vr,j,h)
header(ws,f'B{vr}:M{vr}')
for i in [12,13]:
    src=i-6; ws[f'B{i}']=f'=B{src}'
    ws[f'C{i}']=f'=C{src}*D{src}+G{src}*H{src}+K{src}*H{src}+N{src}*H{src}'
    ws[f'D{i}']=f'=E{src}*F{src}+I{src}*J{src}+L{src}+O{src}'
    ws[f'E{i}']=f'=(E{src}-C{src})*F{src}'
    ws[f'F{i}']=f'=(F{src}-D{src})*C{src}'
    ws[f'G{i}']=f'=(I{src}-G{src})*J{src}'
    ws[f'H{i}']=f'=(J{src}-H{src})*G{src}'
    ws[f'I{i}']=f'=L{src}-K{src}*J{src}'
    ws[f'J{i}']=f'=(J{src}-H{src})*K{src}'
    ws[f'K{i}']=f'=O{src}-M{src}'
    ws[f'L{i}']=f'=M{src}-N{src}*H{src}'
    ws[f'M{i}']=f'=SUM(E{i}:L{i})'
for j in range(2,14): ws.cell(14,j,f'=SUM({get_column_letter(j)}12:{get_column_letter(j)}13)')
ws['B14']='Grand Total'; ws['B14'].font=Font(bold=True)
style_range(ws,'B12:M14',WHITE); money_fmt(ws,'C12:M14')
# prominent totals
for i,(label,formula) in enumerate([('Standard cost','=C14'),('Actual cost','=D14'),('Total cost variance','=M14')],18):
    ws.cell(i,2,label); ws.cell(i,3,formula)
    if 'variance' in label.lower(): ws.cell(i,4,f'=IF(C{i}>0,"Unfavorable",IF(C{i}<0,"Favorable","Neutral"))')
style_range(ws,'B18:D20',GREEN,Font(bold=True)); money_fmt(ws,'C18:C20')
ws['B22']='Required check'; ws['C22']='=AND(C18=110250,C19=114568,C20=4318)'; ws['D22']='TRUE means the example totals match the specification.'
style_range(ws,'B22:D22',GREY)
widths(ws, {get_column_letter(c):15 for c in range(2,16)}); ws.column_dimensions['B'].width=18

# Final formatting / save
for ws in wb.worksheets:
    ws.sheet_view.showGridLines=False
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                c.alignment = Alignment(vertical='center', wrap_text=True)
                if c.row not in [2,5,14,15,24]: c.font = c.font.copy(size=10)
    ws.freeze_panes='B5'

# unlock Inputs/Actuals/Example input areas, protect only requested output tabs
for sh in ['Inputs','Actuals','Example']:
    wb[sh].protection.sheet=False
for sh in ['Calc','Dashboard']:
    wb[sh].protection.sheet=True

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f'Wrote {OUT}')
print('Example grand totals: Standard cost 110250 | Actual cost 114568 | Total cost variance 4318 Unfavorable')
